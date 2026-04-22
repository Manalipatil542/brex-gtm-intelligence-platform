import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import json

#-------------LOAD DATA-------------
# Read the CSV generated 
df = pd.read_csv("leads.csv")

# Read the summary JSON
with open("summary.json") as f:
    s = json.load(f)

# Create a brand new Excel workbook
wb = Workbook()

print("Data loaded successfully!")
print(f"leads loaded: {len(df)} rows")
print(f"Toatl ARR from JSON: ${s['total_arr']:,.0f}")
print(f"Workbook created: {wb}")

# ── COLOR PALETTE ──
# These are hex color codes — same as in design tools like Figma
DARK_BLUE = "0F2D6B"    # dark navy — used for headers
BLUE      = "1B4FD8"    # bright blue — used for section headers
LIGHT     = "EEF2FF"    # very light blue — used for alternating rows
WHITE     = "FFFFFF"    # white
GREEN     = "15803D"    # green — used for positive metrics
RED       = "EF4444"    # red — used for negative metrics
GRAY      = "6B7280"    # gray — used for secondary text
ORANGE    = "D95A1A"    # orange — used for partner metrics

# HELPER FUNCTION 1: HEADER CELL
# This function formats a cell as a dark blue header with white text
def hdr(ws, row, col, value):
    c = ws.cell(row=row, column=col, value=value)
    c.font      = Font(name="Arial", bold=True, color=WHITE, size=10)
    c.fill      = PatternFill("solid", fgColor=DARK_BLUE)
    c.alignment = Alignment(horizontal="center", vertical="center")
    return c

# HELPER FUNCTION 2: REGULAR CELL
# This function writes a value into a cell with basic formatting
def cell(ws, row, col, value, bold=False, color="222222", bg=None, align="left", fmt=None):
    c = ws.cell(row=row, column=col, value=value)
    c.font      = Font(name="Arial", bold=bold, color=color, size=10)
    c.alignment = Alignment(horizontal=align, vertical="center")
    if bg:
        c.fill = PatternFill("solid", fgColor=bg)
    if fmt:
        c.number_format = fmt
    return c

# HELPER FUNCTION 3: BORDER 
# This adds a thin gray border around a range of cells
def add_borders(ws, min_row, min_col, max_row, max_col):
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for row in range(min_row, max_row + 1):
        for col in range(min_col, max_col + 1):
            ws.cell(row, col).border = border

print("Colors and helper functions ready!")

# SHEET 1: EXECUTIVE SUMMARY 
# Get the default sheet and rename it
ws1 = wb.active
ws1.title = "Executive Summary"
ws1.sheet_view.showGridLines = False  # hides the gray grid lines

# TITLE BANNER
# Merge cells A1 to H1 to make one wide title cell
ws1.merge_cells("A1:H1")
title = ws1["A1"]
title.value      = "BREX  ·  GTM Intelligence Platform  ·  2024 Annual Review"
title.font       = Font(name="Arial", bold=True, size=16, color=WHITE)
title.fill       = PatternFill("solid", fgColor=DARK_BLUE)
title.alignment  = Alignment(horizontal="center", vertical="center")
ws1.row_dimensions[1].height = 45

# SUBTITLE 
ws1.merge_cells("A2:H2")
sub = ws1["A2"]
sub.value     = "Prepared by: [Your Name]  |  Target Role: Marketing Ops & BizOps  |  Brex / Capital One"
sub.font      = Font(name="Arial", italic=True, size=10, color=GRAY)
sub.fill      = PatternFill("solid", fgColor=LIGHT)
sub.alignment = Alignment(horizontal="center", vertical="center")
ws1.row_dimensions[2].height = 22

# KPI SECTION HEADER 
ws1.merge_cells("A4:H4")
hdr(ws1, 4, 1, "KEY PERFORMANCE INDICATORS  —  FULL YEAR 2024")
ws1.row_dimensions[4].height = 28

# KPI VALUES 
# List of KPIs — label, value, format
kpis = [
    ("Total Leads",    s["total_leads"],  "#,##0"),
    ("MQLs",           s["total_mqls"],   "#,##0"),
    ("SQLs",           s["total_sqls"],   "#,##0"),
    ("Opportunities",  s["total_opps"],   "#,##0"),
    ("Closed Won",     s["total_won"],    "#,##0"),
    ("Total ARR",      s["total_arr"],    "$#,##0"),
    ("MQL Rate",       s["mql_rate"]/100, "0.0%"),
    ("Win Rate",       s["win_rate"]/100, "0.0%"),
]

# Row 5 = labels, Row 6 = values
for i, (label, value, fmt) in enumerate(kpis):
    col = i + 1
    # Header label
    hdr(ws1, 5, col, label)
    ws1.row_dimensions[5].height = 26
    # Value cell
    c = ws1.cell(row=6, column=col, value=value)
    c.font         = Font(name="Arial", bold=True, size=14, color=DARK_BLUE)
    c.number_format = fmt
    c.alignment    = Alignment(horizontal="center", vertical="center")
    c.fill         = PatternFill("solid", fgColor=LIGHT)
    ws1.row_dimensions[6].height = 36

# SOURCE BREAKDOWN TABLE
ws1.merge_cells("A8:H8")
hdr(ws1, 8, 1, "LEAD SOURCE PERFORMANCE")
ws1.row_dimensions[8].height = 28

# Table headers
src_headers = ["Source", "Total Leads", "MQLs", "Won", "ARR", "MQL Rate", "Win Rate"]
for i, h in enumerate(src_headers):
    hdr(ws1, 9, i+1, h)
ws1.row_dimensions[9].height = 24

# Table rows — one row per source
by_source = df.groupby("source").agg(
    leads = ("lead_id", "count"),
    mqls  = ("is_mql",  "sum"),
    won   = ("is_won",  "sum"),
    arr   = ("arr_usd", "sum")
).reset_index()

for row_i, row in by_source.iterrows():
    r  = row_i + 10
    bg = LIGHT if row_i % 2 == 0 else WHITE

    mql_rate = row["mqls"] / row["leads"] if row["leads"] > 0 else 0
    win_rate = row["won"]  / row["leads"] if row["leads"] > 0 else 0

    cell(ws1, r, 1, row["source"],        bg=bg, align="left")
    cell(ws1, r, 2, row["leads"],         bg=bg, align="right", fmt="#,##0")
    cell(ws1, r, 3, row["mqls"],          bg=bg, align="right", fmt="#,##0")
    cell(ws1, r, 4, row["won"],           bg=bg, align="right", fmt="#,##0")
    cell(ws1, r, 5, row["arr"],           bg=bg, align="right", fmt="$#,##0")
    cell(ws1, r, 6, mql_rate,             bg=bg, align="right", fmt="0.0%")
    cell(ws1, r, 7, win_rate,             bg=bg, align="right", fmt="0.0%")

# Add borders around the whole table
add_borders(ws1, 9, 1, 9 + len(by_source), 7)

# COLUMN WIDTHS 
ws1.column_dimensions["A"].width = 22
ws1.column_dimensions["B"].width = 13
ws1.column_dimensions["C"].width = 10
ws1.column_dimensions["D"].width = 10
ws1.column_dimensions["E"].width = 14
ws1.column_dimensions["F"].width = 11
ws1.column_dimensions["G"].width = 11

print("Sheet 1 - Executive Summary built!")

# SHEET 2: PIPELINE TREND 
ws2 = wb.create_sheet("Pipeline Trend")
ws2.sheet_view.showGridLines = False

# TITLE
ws2.merge_cells("A1:J1")
t = ws2["A1"]
t.value     = "MONTHLY PIPELINE TREND  —  JAN 2024 TO DEC 2024"
t.font      = Font(name="Arial", bold=True, size=14, color=WHITE)
t.fill      = PatternFill("solid", fgColor=DARK_BLUE)
t.alignment = Alignment(horizontal="center", vertical="center")
ws2.row_dimensions[1].height = 36

# TABLE HEADERS 
trend_headers = [
    "Month", "New Leads", "MQLs", "SQLs",
    "Opportunities", "Closed Won",
    "Pipeline ARR", "Closed ARR",
    "Avg Deal Size", "Sales Cycle (days)"
]
for i, h in enumerate(trend_headers):
    hdr(ws2, 2, i+1, h)
ws2.row_dimensions[2].height = 24

# GENERATE MONTHLY DATA 
# simulate 12 months of growing pipeline
# Each month grows slightly — this shows positive business momentum

import random as rnd

months = [
    "Jan 2024", "Feb 2024", "Mar 2024", "Apr 2024",
    "May 2024", "Jun 2024", "Jul 2024", "Aug 2024",
    "Sep 2024", "Oct 2024", "Nov 2024", "Dec 2024"
]

for i, month in enumerate(months):
    r  = i + 3
    bg = LIGHT if i % 2 == 0 else WHITE

    # Each metric grows slightly each month
    growth = 1 + (i * 0.04)

    new_leads  = int(35  * growth)
    mqls       = int(10  * growth)
    sqls       = int(5   * growth)
    opps       = int(3   * growth)
    won        = int(1.2 * growth)
    pipeline   = int(320000 * growth)
    closed     = int(85000  * growth)
    deal_size  = int(38000  * growth)
    cycle_days = int(65 - i * 0.8)

    cell(ws2, r, 1,  month,      bg=bg, align="left")
    cell(ws2, r, 2,  new_leads,  bg=bg, align="right", fmt="#,##0")
    cell(ws2, r, 3,  mqls,       bg=bg, align="right", fmt="#,##0")
    cell(ws2, r, 4,  sqls,       bg=bg, align="right", fmt="#,##0")
    cell(ws2, r, 5,  opps,       bg=bg, align="right", fmt="#,##0")
    cell(ws2, r, 6,  won,        bg=bg, align="right", fmt="#,##0")
    cell(ws2, r, 7,  pipeline,   bg=bg, align="right", fmt="$#,##0")
    cell(ws2, r, 8,  closed,     bg=bg, align="right", fmt="$#,##0")
    cell(ws2, r, 9,  deal_size,  bg=bg, align="right", fmt="$#,##0")
    cell(ws2, r, 10, cycle_days, bg=bg, align="right")

# Add borders
add_borders(ws2, 2, 1, 14, 10)

# COLUMN WIDTHS 
ws2.column_dimensions["A"].width = 14
ws2.column_dimensions["B"].width = 12
ws2.column_dimensions["C"].width = 10
ws2.column_dimensions["D"].width = 10
ws2.column_dimensions["E"].width = 14
ws2.column_dimensions["F"].width = 12
ws2.column_dimensions["G"].width = 15
ws2.column_dimensions["H"].width = 14
ws2.column_dimensions["I"].width = 14
ws2.column_dimensions["J"].width = 18

print("Sheet 2 - Pipeline Trend built!")


# SHEET 3: PARTNER TRACKER 
ws3 = wb.create_sheet("Partner Tracker")
ws3.sheet_view.showGridLines = False

# TITLE
ws3.merge_cells("A1:L1")
t = ws3["A1"]
t.value     = "PARTNER OPERATIONS TRACKER  —  BREX BD TEAM"
t.font      = Font(name="Arial", bold=True, size=14, color=WHITE)
t.fill      = PatternFill("solid", fgColor=DARK_BLUE)
t.alignment = Alignment(horizontal="center", vertical="center")
ws3.row_dimensions[1].height = 36

# TABLE HEADERS 
partner_headers = [
    "Partner Name", "Type", "Activation Date",
    "Deals Sourced", "Deals Won", "Win Rate",
    "Partner ARR", "Engagement Score", "NRR %",
    "QBR Done", "Status", "Co-Marketing"
]
for i, h in enumerate(partner_headers):
    hdr(ws3, 2, i+1, h)
ws3.row_dimensions[2].height = 24

# PARTNER DATA 
# 7 realistic Brex partners with different health statuses
partners = [
    ("Fifth Third Bank",  "Bank",        "2023-03-15", 38, 22, 58, 1240000, 88, 138, "Yes", "Active",  "Active"),
    ("Oracle Cloud",      "Cloud ERP",   "2023-06-01", 29, 15, 52, 980000,  82, 125, "Yes", "Active",  "Active"),
    ("DoorDash Business", "Travel",      "2023-09-10", 27, 12, 44, 760000,  74, 108, "Yes", "Active",  "Inactive"),
    ("Navan",             "Travel",      "2023-11-20", 22, 13, 61, 680000,  91, 143, "Yes", "Active",  "Active"),
    ("Zip Procurement",   "Procurement", "2024-01-05", 31, 11, 35, 920000,  67, 95,  "No",  "At Risk", "Inactive"),
    ("Anthropic",         "AI/ML",       "2024-03-01", 19, 9,  48, 580000,  79, 119, "Yes", "Active",  "Active"),
    ("Wiz Security",      "Security",    "2024-05-15", 16, 5,  29, 352347,  55, 91,  "No",  "At Risk", "Inactive"),
]

for row_i, partner in enumerate(partners):
    r  = row_i + 3
    bg = LIGHT if row_i % 2 == 0 else WHITE

    # Unpack all partner values
    name, ptype, date, deals, won, wr, arr, eng, nrr, qbr, status, comktg = partner

    # Status color — green for Active, orange for At Risk
    status_color = GREEN if status == "Active" else "F59E0B"

    cell(ws3, r, 1,  name,        bg=bg, align="left")
    cell(ws3, r, 2,  ptype,       bg=bg, align="left")
    cell(ws3, r, 3,  date,        bg=bg, align="center")
    cell(ws3, r, 4,  deals,       bg=bg, align="right",  fmt="#,##0")
    cell(ws3, r, 5,  won,         bg=bg, align="right",  fmt="#,##0")
    cell(ws3, r, 6,  wr/100,      bg=bg, align="right",  fmt="0.0%")
    cell(ws3, r, 7,  arr,         bg=bg, align="right",  fmt="$#,##0")
    cell(ws3, r, 8,  eng,         bg=bg, align="right",  fmt="0.0")
    cell(ws3, r, 9,  nrr/100,     bg=bg, align="right",  fmt="0.0%")
    cell(ws3, r, 10, qbr,         bg=bg, align="center")

    # Status cell gets special color
    c = cell(ws3, r, 11, status,  bg=bg, align="center")
    c.font = Font(name="Arial", bold=True, color=status_color, size=10)

    cell(ws3, r, 12, comktg,      bg=bg, align="center")

# Add borders
add_borders(ws3, 2, 1, 2 + len(partners), 12)

# COLUMN WIDTHS 
ws3.column_dimensions["A"].width = 20
ws3.column_dimensions["B"].width = 14
ws3.column_dimensions["C"].width = 16
ws3.column_dimensions["D"].width = 14
ws3.column_dimensions["E"].width = 12
ws3.column_dimensions["F"].width = 11
ws3.column_dimensions["G"].width = 14
ws3.column_dimensions["H"].width = 17
ws3.column_dimensions["I"].width = 10
ws3.column_dimensions["J"].width = 11
ws3.column_dimensions["K"].width = 11
ws3.column_dimensions["L"].width = 13

print("Sheet 3 - Partner Tracker built!")


# SHEET 4: LEAD SCORING MODEL 
ws4 = wb.create_sheet("Lead Scoring Model")
ws4.sheet_view.showGridLines = False

# TITLE 
ws4.merge_cells("A1:E1")
t = ws4["A1"]
t.value     = "LEAD SCORING FRAMEWORK  —  BREX B2B FUNNEL"
t.font      = Font(name="Arial", bold=True, size=14, color=WHITE)
t.fill      = PatternFill("solid", fgColor=DARK_BLUE)
t.alignment = Alignment(horizontal="center", vertical="center")
ws4.row_dimensions[1].height = 36

# SECTION 1: FIRMOGRAPHIC SCORING 
ws4.merge_cells("A3:E3")
hdr(ws4, 3, 1, "FIRMOGRAPHIC SIGNALS — Company Fit Score")
ws4.row_dimensions[3].height = 28

firm_headers = ["Category", "Criteria", "Points", "Rationale"]
for i, h in enumerate(firm_headers):
    hdr(ws4, 4, i+1, h)
ws4.row_dimensions[4].height = 24

firm_data = [
    ("Company Size",    "Enterprise (500+ emp)",   30, "Highest ACV potential"),
    ("Company Size",    "Mid-Market (50-500)",      20, "Core Brex ICP"),
    ("Company Size",    "Startup (<50 emp)",        10, "Volume play"),
    ("Funding Stage",   "Series B / C",             25, "High spending velocity"),
    ("Funding Stage",   "Series A",                 18, "Growing fast"),
    ("Funding Stage",   "Enterprise",               30, "Largest deals"),
    ("Industry",        "FinTech / SaaS",           20, "Highest product fit"),
    ("Industry",        "E-Commerce",               15, "Heavy card spend"),
    ("Industry",        "Other Tech",               10, "Moderate fit"),
]

for row_i, row in enumerate(firm_data):
    r  = row_i + 5
    bg = LIGHT if row_i % 2 == 0 else WHITE
    cell(ws4, r, 1, row[0], bg=bg, align="left")
    cell(ws4, r, 2, row[1], bg=bg, align="left")
    cell(ws4, r, 3, row[2], bg=bg, align="right", fmt="#,##0")
    cell(ws4, r, 4, row[3], bg=bg, align="left")

add_borders(ws4, 4, 1, 4 + len(firm_data), 4)

# SECTION 2: BEHAVIORAL SCORING 
ws4.merge_cells("A16:E16")
hdr(ws4, 16, 1, "BEHAVIORAL SIGNALS — Intent Score")
ws4.row_dimensions[16].height = 28

for i, h in enumerate(firm_headers):
    hdr(ws4, 17, i+1, h)
ws4.row_dimensions[17].height = 24

behav_data = [
    ("Demo Request",       "Requested live demo",           40, "Strongest intent signal"),
    ("Partner Referral",   "Referred by active partner",    35, "Trust pre-built"),
    ("Chatbot Engaged",    "Used Qualified on website",     30, "Active interest"),
    ("Pricing Page",       "Visited pricing page 2+ times", 25, "Evaluation mode"),
    ("Content Download",   "Downloaded ROI calculator",     20, "Self-educating"),
    ("Event Attendance",   "Attended Brex webinar",         20, "High intent"),
    ("Email Engagement",   "Opened 3+ emails in sequence",  15, "Nurture responsive"),
]

for row_i, row in enumerate(behav_data):
    r  = row_i + 18
    bg = LIGHT if row_i % 2 == 0 else WHITE
    cell(ws4, r, 1, row[0], bg=bg, align="left")
    cell(ws4, r, 2, row[1], bg=bg, align="left")
    cell(ws4, r, 3, row[2], bg=bg, align="right", fmt="#,##0")
    cell(ws4, r, 4, row[3], bg=bg, align="left")

add_borders(ws4, 17, 1, 17 + len(behav_data), 4)

# SECTION 3: ROUTING THRESHOLDS 
ws4.merge_cells("A27:E27")
hdr(ws4, 27, 1, "ROUTING LOGIC — What Happens at Each Score")
ws4.row_dimensions[27].height = 28

thresh_headers = ["Score Range", "Classification", "Action", "Tool Used"]
for i, h in enumerate(thresh_headers):
    hdr(ws4, 28, i+1, h)
ws4.row_dimensions[28].height = 24

thresh_data = [
    ("80 – 100", "Hot MQL",  "Route instantly to AE",          GREEN),
    ("50 – 79",  "MQL",      "Enroll in SDR Outreach cadence", BLUE),
    ("30 – 49",  "Nurture",  "Marketo 30-day drip program",    ORANGE),
    ("0 – 29",   "Cold",     "Monitor only — no outreach",     GRAY),
]

tools = ["LeanData → Salesforce", "Outreach → SDR", "Marketo", "Segment"]

for row_i, (score, cls, action, color) in enumerate(thresh_data):
    r  = row_i + 29
    bg = LIGHT if row_i % 2 == 0 else WHITE

    cell(ws4, r, 1, score,        bg=bg, align="center")
    # Classification gets its own color
    c = cell(ws4, r, 2, cls,      bg=bg, align="center")
    c.font = Font(name="Arial", bold=True, color=color, size=10)
    cell(ws4, r, 3, action,       bg=bg, align="left")
    cell(ws4, r, 4, tools[row_i], bg=bg, align="left")

add_borders(ws4, 28, 1, 28 + len(thresh_data), 4)

# COLUMN WIDTHS 
ws4.column_dimensions["A"].width = 20
ws4.column_dimensions["B"].width = 30
ws4.column_dimensions["C"].width = 10
ws4.column_dimensions["D"].width = 30

print("Sheet 4 - Lead Scoring Model built!")

# SAVE THE WORKBOOK 
wb.save("Brex_GTM_Model.xlsx")

print("")
print("=" * 50)
print("EXCEL FILE SAVED SUCCESSFULLY!")
print("=" * 50)
print("")
print("File: Brex_GTM_Model.xlsx")
print("Location: brex_project folder on your Desktop")
print("")
print("Sheets built:")
print("Sheet 1 — Executive Summary")
print("Sheet 2 — Pipeline Trend")
print("Sheet 3 — Partner Tracker")
print("Sheet 4 — Lead Scoring Model")
print("")
print("Open it in Excel now!")